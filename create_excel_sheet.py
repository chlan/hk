#!/usr/bin/env python

import openpyxl as xl

pl = [['Sasse', 'Jonas'],
      ['Sayman', 'Savas'],
      ['Schliemann', 'Tobias'],
      ['Schliewack', 'Leon'],
      ['Schlüter', 'Simon'],
      ['Schmidt', 'Leon Marc'],
     ]

print(pl)

for r in range(0, len(pl)):
    print(pl[r])
    n0= pl[r][0] + ', ' + pl[r][1]
    n1= pl[r][0] + '_' + pl[r][1]
    fn1= 'bewertung_' + n1 + '_CHL' + '.xlsx'
    fn2= 'bewertung_' + n1 + '.xlsx'
    fn3= 'fragen_' + n1 + '.xlsx'
    print(r,n0, n1, fn1, fn2)
    wb1= xl.load_workbook('template_bewertung.xlsx')
    wb2= xl.load_workbook('template_fragen.xlsx')
    sheet= wb1['Dokumentation']
    cell= sheet.cell(7,2)
    cell.value= n0
    wb1.save(fn1)
    wb1.save(fn2)
    sheet= wb2['Tabelle1']
    cell= sheet.cell(1,3)
    cell.value= n0
    wb2.save(fn3)

